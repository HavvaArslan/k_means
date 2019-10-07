# -*- coding: utf-8 -*-
"""
@author: HavvaArslan
"""
import numpy as np
import pandas as pd
import xlrd
from openpyxl import load_workbook

k = input("k değeri gir   ")
veriler=[]
kume_degiskenleri=['C1','C2']
kumeler=[]
guncellenen_kumeler=[]
merkez_degerleri=[]
kume_ici_degisim=[]
count=0
uzaklik=[]
skor=1


#%%
def veri_okuma():
    a = xlrd.open_workbook("k-means.xlsx")
    table = a.sheet_by_name(u'Sayfa1')
    sayac=0
    global count
    for sayfa in a.sheets():
        for satir in range(sayfa.nrows):
            if sayac!=0:    
                veriler.append(table.row_values(satir))
            sayac+=1
            count=sayac
            for sutun in range(sayfa.ncols):
                degerler=[]
                degerler.append(sayfa.cell(satir, sutun).value)
        

#%%  
def veri_yazma(kumeler):
    wb =load_workbook('k-means.xlsx')
    sheet=wb.worksheets[0]
    row_count = int(sheet.max_row)
    ws = wb.active
    
    for i in range(row_count-1):
        ws['D'+str(i+2)]=kumeler[i]
    wb.save("k-means.xlsx")

#%% 
def kume_olustur():
    for i in range(count-1):
        degisken="C"+str(np.random.randint(1,int(k)+1))
        kumeler.append(degisken);
        print("random",degisken)
        
#%%
def merkez_hesapla():
    merkez1=0
    merkez2=0
    
    for j in range(len(kume_degiskenleri)):
        sayac=0
        for i in range(len(veriler)):
            if veriler[i][3] == kume_degiskenleri[j]:
                sayac+=1;
                merkez1+=veriler[i][1]
                merkez2+=veriler[i][2]
        merkez_degerleri.append([merkez1/sayac , merkez2/sayac])
        merkez1=0
        merkez2=0

#%%
def kume_ici_degisim_hesapla():
    e=0
    
    for j in range(len(kume_degiskenleri)):
        for i in range(len(veriler)):
            if veriler[i][3] == kume_degiskenleri[j]:
                e+=(np.math.pow(veriler[i][1]-merkez_degerleri[j][0],2))+(np.math.pow(veriler[i][2]-merkez_degerleri[j][1],2))
        kume_ici_degisim.append(e)
        e=0
       
#%%
def toplam_kare_hata_hesapla():
    E=0
    for i in range(len(kume_ici_degisim)):
        E+=kume_ici_degisim[i]
    
    return E
        
        
        
#%%
def merkez_uzaklik_hesapla():
    
   for i in range(len(merkez_degerleri)):
       dizi=[]
       for j in range(len(veriler)):
           dizi.append(np.math.sqrt(np.math.pow(merkez_degerleri[i][0]-
                  veriler[j][1],2)+np.math.pow(merkez_degerleri[i][1]-
                         veriler[j][2],2)))
       uzaklik.append(dizi)
    
   for i in range(len(uzaklik[0])):
       if uzaklik[0][i] < uzaklik[1][i]:
           guncellenen_kumeler.append("C1")
       else:
           guncellenen_kumeler.append("C2")
           
    
                
#%%
veri_okuma()
kume_olustur()
veri_yazma(kumeler)
merkez_hesapla()
kume_ici_degisim_hesapla()
sonuc=toplam_kare_hata_hesapla()
merkez_uzaklik_hesapla()

while skor==1:
    if guncellenen_kumeler==kumeler:
        skor=0
        break
    else:
        kumeler=guncellenen_kumeler
        veri_yazma(guncellenen_kumeler)
        veri_okuma()
        merkez_hesapla()
        kume_ici_degisim_hesapla()
        sonuc=toplam_kare_hata_hesapla()
        merkez_uzaklik_hesapla()
          
print("Bitti :)")

       
